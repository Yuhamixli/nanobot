"""
Local knowledge store: chunk documents, embed with BGE-small-zh, store in ChromaDB.

Requires: pip install nanobot-ai[rag]
"""

import hashlib
import os
import time
from pathlib import Path
from typing import Any

if "HF_ENDPOINT" not in os.environ:
    os.environ["HF_ENDPOINT"] = "https://hf-mirror.com"
# Approximate tokens to chars for Chinese (BGE/sentence-transformers)
CHARS_PER_TOKEN = 2
COLLECTION_NAME = "nanobot_kb"
WEB_CACHE_COLLECTION = "nanobot_kb_web_cache"
# 短期知识目录：爬取内容、web cache，按 TTL 定期清理
SHORT_TERM_DIR = "短期"
WEB_CACHE_DIR = f"{SHORT_TERM_DIR}/_cache_web"  # web 缓存置于短期下
CLEANUP_INTERVAL_DAYS = 7
LONG_TERM_DIR = "长期"  # 长期知识：制度、手册，不自动清理
SUPPORTED_EXTENSIONS = {".txt", ".md", ".pdf", ".docx", ".xlsx"}


def _chunk_text(text: str, chunk_size: int, overlap: int) -> list[str]:
    """Split text into overlapping chunks (sizes in tokens, converted to chars)."""
    if not text.strip():
        return []
    size_chars = chunk_size * CHARS_PER_TOKEN
    overlap_chars = overlap * CHARS_PER_TOKEN
    step = max(1, size_chars - overlap_chars)
    chunks = []
    start = 0
    while start < len(text):
        end = min(start + size_chars, len(text))
        chunk = text[start:end].strip()
        if chunk:
            chunks.append(chunk)
        start += step
    return chunks


def _load_text(path: Path) -> str:
    """Load plain text or markdown."""
    return path.read_text(encoding="utf-8", errors="replace")


def _load_pdf(path: Path) -> str:
    """Load PDF text via pypdf."""
    from pypdf import PdfReader
    reader = PdfReader(path)
    parts = []
    for page in reader.pages:
        parts.append(page.extract_text() or "")
    return "\n\n".join(parts)


def _load_docx(path: Path) -> str:
    """Load Word document text."""
    from docx import Document
    doc = Document(path)
    return "\n\n".join(p.text for p in doc.paragraphs if p.text.strip())


def _load_xlsx(path: Path) -> str:
    """Load Excel sheet text (all sheets, cell values joined)."""
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True, data_only=True)
    parts = []
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            parts.append(" ".join(str(c) if c is not None else "" for c in row))
    return "\n\n".join(parts)


def _load_document(path: Path) -> str:
    """Load document content by extension."""
    path = path.resolve()
    if not path.exists():
        raise FileNotFoundError(str(path))
    suffix = path.suffix.lower()
    if suffix in (".txt", ".md"):
        return _load_text(path)
    if suffix == ".pdf":
        return _load_pdf(path)
    if suffix == ".docx":
        return _load_docx(path)
    if suffix == ".xlsx":
        return _load_xlsx(path)
    raise ValueError(f"Unsupported format: {suffix}. Supported: {SUPPORTED_EXTENSIONS}")


def get_rag_import_error() -> str | None:
    """
    Return None if RAG deps are OK, else a short message (e.g. "chromadb" or "sentence_transformers").
    """
    try:
        import chromadb
    except ImportError as e:
        return f"chromadb ({e!s})"
    try:
        from sentence_transformers import SentenceTransformer
    except ImportError as e:
        return f"sentence_transformers ({e!s})"
    return None


def get_store(
    workspace: Path,
    chunk_size: int = 512,
    chunk_overlap: int = 200,
    top_k: int = 5,
) -> "KnowledgeStore | None":
    """
    Return a KnowledgeStore if RAG dependencies are installed, else None.
    """
    if get_rag_import_error() is not None:
        return None
    return KnowledgeStore(
        workspace=Path(workspace),
        chunk_size=chunk_size,
        chunk_overlap=chunk_overlap,
        top_k=top_k,
    )


class KnowledgeStore:
    """
    Chunk documents, embed with BGE-small-zh, store and query via ChromaDB.
    Data is stored under workspace/knowledge_db.
    """

    def __init__(
        self,
        workspace: Path,
        chunk_size: int = 512,
        chunk_overlap: int = 200,
        top_k: int = 5,
    ):
        self.workspace = Path(workspace)
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
        self.top_k = top_k
        self._db_path = self.workspace / "knowledge_db"
        self._db_path.mkdir(parents=True, exist_ok=True)
        self._client = None
        self._collection = None
        self._web_cache_collection = None
        self._model = None

    def _get_client(self):
        import chromadb
        from chromadb.config import Settings
        if self._client is None:
            self._client = chromadb.PersistentClient(
                path=str(self._db_path),
                settings=Settings(anonymized_telemetry=False),
            )
        return self._client

    def _get_collection(self):
        if self._collection is None:
            client = self._get_client()
            # hnsw:search_ef 提高检索召回，避免 n_results 较小时漏掉相关文档
            self._collection = client.get_or_create_collection(
                name=COLLECTION_NAME,
                metadata={
                    "description": "nanobot local knowledge base",
                    "hnsw:search_ef": 64,
                },
            )
        return self._collection

    def _get_web_cache_collection(self):
        """Get or create the web cache collection (separate from main KB)."""
        if self._web_cache_collection is None:
            client = self._get_client()
            self._web_cache_collection = client.get_or_create_collection(
                name=WEB_CACHE_COLLECTION,
                metadata={"description": "nanobot web search cache", "hnsw:search_ef": 64},
            )
        return self._web_cache_collection

    def _get_model(self):
        if self._model is None:
            import os
            # 未设置时使用国内镜像，避免直连 huggingface.co 超时；国外用户可设 HF_ENDPOINT=https://huggingface.co
            if "HF_ENDPOINT" not in os.environ:
                os.environ["HF_ENDPOINT"] = "https://hf-mirror.com"
            self._model = __import__("sentence_transformers").SentenceTransformer(
                "BAAI/bge-small-zh-v1.5"
            )
        return self._model

    def _embed(self, texts: list[str]) -> list[list[float]]:
        model = self._get_model()
        embeddings = model.encode(texts, normalize_embeddings=True)
        return embeddings.tolist()

    def add_documents(
        self,
        paths: list[Path],
        *,
        skip_unsupported: bool = True,
    ) -> dict[str, Any]:
        """
        Ingest files/directories: load, chunk, embed, add to ChromaDB.
        Returns summary: added_count, skipped, errors.
        """
        added = 0
        skipped: list[str] = []
        errors: list[str] = []

        all_chunks: list[str] = []
        all_metadatas: list[dict] = []
        all_ids: list[str] = []

        for p in paths:
            path = Path(p).resolve()
            if not path.exists():
                errors.append(f"Not found: {path}")
                continue
            if path.is_file():
                to_process = [path]
            else:
                to_process = [
                    f for f in path.rglob("*")
                    if f.is_file() and f.suffix.lower() in SUPPORTED_EXTENSIONS
                ]
            for fp in to_process:
                try:
                    text = _load_document(fp)
                except Exception as e:
                    errors.append(f"{fp}: {e}")
                    continue
                if not text.strip():
                    skipped.append(str(fp))
                    continue
                chunks = _chunk_text(text, self.chunk_size, self.chunk_overlap)
                rel = fp.relative_to(path if path.is_dir() else path.parent)
                for i, c in enumerate(chunks):
                    doc_id = f"{rel!s}_{i}".replace("\\", "/")
                    all_ids.append(doc_id)
                    all_chunks.append(c)
                    all_metadatas.append({"source": str(rel), "chunk": i})

        if not all_chunks:
            return {"added": 0, "skipped": skipped, "errors": errors}

        coll = self._get_collection()
        embeddings = self._embed(all_chunks)
        coll.add(
            ids=all_ids,
            documents=all_chunks,
            embeddings=embeddings,
            metadatas=all_metadatas,
        )
        added = len(all_chunks)
        return {"added": added, "skipped": skipped, "errors": errors}

    def add_to_web_cache(
        self,
        text: str,
        query: str = "",
        url: str = "",
        tool_name: str = "",
    ) -> None:
        """
        Save web search/fetch result to workspace/knowledge/短期/_cache_web/ and ingest into web cache collection.
        """
        if not text or not text.strip():
            return
        cache_dir = self.workspace / "knowledge" / WEB_CACHE_DIR
        cache_dir.mkdir(parents=True, exist_ok=True)
        ts = int(time.time())
        # Use hash of query+url to avoid filename collision
        key = f"{query or url}_{ts}"
        h = hashlib.md5(key.encode()).hexdigest()[:12]
        fname = f"web_{ts}_{h}.md"
        path = cache_dir / fname
        header = f"---\nquery: {query}\nurl: {url}\ntool: {tool_name}\ntimestamp: {ts}\n---\n\n"
        path.write_text(header + text.strip(), encoding="utf-8")
        try:
            chunks = _chunk_text(text, self.chunk_size, self.chunk_overlap)
            if not chunks:
                return
            rel = f"{WEB_CACHE_DIR}/{fname}"
            all_ids = [f"{rel}_{i}" for i in range(len(chunks))]
            all_metadatas = [{"source": rel, "chunk": i} for i in range(len(chunks))]
            embeddings = self._embed(chunks)
            wc = self._get_web_cache_collection()
            wc.add(ids=all_ids, documents=chunks, embeddings=embeddings, metadatas=all_metadatas)
        except Exception:
            pass  # Don't fail main flow on cache write

    def clear_web_cache(self) -> None:
        """Delete web cache files and clear the web cache collection."""
        cache_dir = self.workspace / "knowledge" / WEB_CACHE_DIR
        if cache_dir.exists():
            for f in cache_dir.glob("*.md"):
                try:
                    f.unlink()
                except OSError:
                    pass
        # 兼容旧版：若存在 knowledge/_cache_web，一并清理
        legacy_cache = self.workspace / "knowledge" / "_cache_web"
        if legacy_cache.exists():
            for f in legacy_cache.glob("*.md"):
                try:
                    f.unlink()
                except OSError:
                    pass
        try:
            client = self._get_client()
            client.delete_collection(WEB_CACHE_COLLECTION)
            self._web_cache_collection = None
        except Exception:
            pass
        # Update last cleanup timestamp
        marker = self.workspace / "knowledge" / WEB_CACHE_DIR / ".last_cleanup"
        marker.parent.mkdir(parents=True, exist_ok=True)
        marker.write_text(str(int(time.time())))

    def should_clear_web_cache(self) -> bool:
        """Check if a week has passed since last cleanup."""
        marker = self.workspace / "knowledge" / WEB_CACHE_DIR / ".last_cleanup"
        if not marker.exists():
            return True
        try:
            last = int(marker.read_text())
            return (time.time() - last) >= CLEANUP_INTERVAL_DAYS * 86400
        except Exception:
            return True

    def cleanup_short_term(self, retention_days: int = 7) -> int:
        """
        清理 短期/ 目录下超期文件及其向量。
        排除 _cache_web（由 clear_web_cache 单独处理）。
        返回删除的文件数。
        """
        short_dir = self.workspace / "knowledge" / SHORT_TERM_DIR
        if not short_dir.exists():
            return 0
        cutoff = time.time() - retention_days * 86400
        deleted = 0
        try:
            coll = self._get_collection()
            for fp in short_dir.rglob("*"):
                if not fp.is_file():
                    continue
                # 跳过 _cache_web（由 clear_web_cache 处理）
                try:
                    rel = fp.relative_to(short_dir)
                    if str(rel).startswith("_cache_web"):
                        continue
                except ValueError:
                    continue
                if fp.stat().st_mtime >= cutoff:
                    continue
                source_key = f"{SHORT_TERM_DIR}/{rel!s}".replace("\\", "/")
                try:
                    data = coll.get(where={"source": source_key}, include=[])
                    ids = data.get("ids", [[]])
                    if ids and ids[0]:
                        coll.delete(ids=ids[0])
                        deleted += 1
                except Exception:
                    pass
                try:
                    fp.unlink()
                except OSError:
                    pass
        except Exception:
            pass
        return deleted

    def search(self, query: str, top_k: int | None = None) -> list[dict[str, Any]]:
        """Return top-k most relevant chunks with content and source. Merges main KB and web cache."""
        k = top_k if top_k is not None else self.top_k
        out: list[dict[str, Any]] = []
        q_emb = self._embed([query])
        # Main collection
        coll = self._get_collection()
        n = coll.count()
        if n > 0:
            main_results = coll.query(
                query_embeddings=q_emb,
                n_results=min(k, n),
                include=["documents", "metadatas", "distances"],
            )
            if main_results and main_results["ids"] and main_results["ids"][0]:
                for i, doc_id in enumerate(main_results["ids"][0]):
                    meta = (main_results["metadatas"][0] or [{}])[i]
                    dist = (main_results["distances"][0] or [0])[i]
                    doc = (main_results["documents"][0] or [""])[i]
                    out.append({
                        "content": doc,
                        "source": meta.get("source", ""),
                        "chunk": meta.get("chunk", 0),
                        "distance": float(dist),
                    })
        # Web cache collection (merge if exists)
        try:
            wc = self._get_web_cache_collection()
            n_wc = wc.count()
            if n_wc > 0:
                wc_results = wc.query(
                    query_embeddings=q_emb,
                    n_results=min(k, n_wc),
                    include=["documents", "metadatas", "distances"],
                )
                if wc_results and wc_results["ids"] and wc_results["ids"][0]:
                    for i, doc_id in enumerate(wc_results["ids"][0]):
                        meta = (wc_results["metadatas"][0] or [{}])[i]
                        dist = (wc_results["distances"][0] or [0])[i]
                        doc = (wc_results["documents"][0] or [""])[i]
                        out.append({
                            "content": doc,
                            "source": meta.get("source", ""),
                            "chunk": meta.get("chunk", 0),
                            "distance": float(dist),
                        })
        except Exception:
            pass
        out.sort(key=lambda x: x.get("distance", 999))
        return out[:k]

    def count(self) -> int:
        """Total number of chunks (main KB + web cache)."""
        try:
            n = self._get_collection().count()
            try:
                n += self._get_web_cache_collection().count()
            except Exception:
                pass
            return n
        except Exception:
            return 0

    def list_sources(self) -> list[str]:
        """Return unique source file names in the knowledge base (main + web cache)."""
        sources: set[str] = set()
        try:
            coll = self._get_collection()
            n = coll.count()
            if n > 0:
                # Limit to avoid loading huge collections; 5k chunks ~= hundreds of files
                data = coll.get(include=["metadatas"], limit=min(n, 5000))
                for m in (data.get("metadatas") or []):
                    if m and isinstance(m, dict) and "source" in m:
                        sources.add(str(m["source"]))
            wc = self._get_web_cache_collection()
            n_wc = wc.count()
            if n_wc > 0:
                data = wc.get(include=["metadatas"], limit=min(n_wc, 2000))
                for m in (data.get("metadatas") or []):
                    if m and isinstance(m, dict) and "source" in m:
                        sources.add(str(m["source"]))
        except Exception:
            pass
        return sorted(sources)
